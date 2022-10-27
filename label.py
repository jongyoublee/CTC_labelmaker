''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
# Release Date : 2022-10-26
# Version : 1.0.0 (Initial Release)
#  
# Make Tube and Box label for Clinical Trial 
# 
#
#
#
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

from openpyxl import Workbook

class LabelMaker():
    #Default Global variable
    column = ['A', 'B', 'C', 'D']
    periodList = ["(1기)", "(2기)", "(3기)", "(4기)", "(5기)", "(6기)"]
    boxPeriodList = ["1기)", "2기)", "3기)", "4기)", "5기)", "6기)"]

    tubeFirstRow = ("시간", "a", "F5", "성분명(P7)")
    boxFirstRow = ("시간", "날짜", "제약회사", "상품명")

    boxList = ["(분석용)", "(보관용)"]

    def __init__(self, SPONSOR, STUDYNAME, SUBJECTS, SEQUENCE, PERIOD, DOSEDAYS, TIMETABLE, FILEPATH):
        self.excel = Workbook()

        self.sponsor = SPONSOR
        self.studyName = STUDYNAME
        self.subjects = SUBJECTS
        self.sequence = SEQUENCE
        self.period = PERIOD
        self.doseDays = DOSEDAYS
        self.timeTable = TIMETABLE
        self.exportFile = FILEPATH

        for i in self.excel.sheetnames:
            self.excel.remove(self.excel[i])

        labelSheet = self.excel.create_sheet("라벨")
        boxSheet = self.excel.create_sheet("박스")
        for col in self.column:
            labelSheet.column_dimensions[col].width = 20
            boxSheet.column_dimensions[col].width = 20

    def makeTubeLabel(self):
        tubeSheet = self.excel["라벨"]
        tubeSheet.append(self.tubeFirstRow)

        LabelData = []

        for pe in range(self.period):
            #print(period[pe])
            if(self.sequence == 1):
                for time in self.timeTable:
                    #print(time)
                    for subject in range(1, self.subjects + 1):
                        #print(subject)
                        tempData = ("(BM)" + time, self.periodList[pe] + " R" + str(self.sequence) + str(subject).zfill(2)  , self.sponsor, self.studyName)
                        print(tempData)
                        LabelData.append(tempData)
            else:
                for time in self.timeTable:
                #print(time)
                    for sq in range(1, self.sequence + 1):
                        for subject in range(1, (self.subjects // self.period) + 1):
                            #print(subject)
                            tempData = ("(BM)" + time, self.periodList[pe] + " R" + str(sq) + str(subject).zfill(2)  , self.sponsor, self.studyName)
                            print(tempData)
                            LabelData.append(tempData)

        for row in LabelData:
            tubeSheet.append(row)

        #self.excel.save(self.exportFile)

    def makeBoxLabel(self):
        boxSheet = self.excel["박스"]
        boxSheet.append(self.boxFirstRow)

        boxData = []

        for pe in range(self.period):
            for time in self.timeTable:
                for box in self.boxList:
                    tempData = (time + box, self.boxPeriodList[pe] + self.doseDays[pe], self.sponsor, self.studyName)
                    print(tempData)
                    boxData.append(tempData)

        for row in boxData:
            boxSheet.append(row)
        

    def saveExcelFile(self):
        self.excel.save(self.exportFile)
'''
gi = 2






#for label in labelData:
    #print(label)


    def initWorkBook(SPONSOR, STUDYNAME, SUBJECTS, SEQUENCE, PERIOD, TIMETABLE, FILEPATH):
        sponsor = SPONSOR
        studyName = STUDYNAME
        subjects = SUBJECTS
        sequence = SEQUENCE
        timeTable = TIMETABLE

        for i in excel.sheetnames:
            excel.remove(excel[i])

        labelSheet = excel.create_sheet("라벨")
        boxSheet = excel.create_sheet("박스")
        for col in column:
            labelSheet.column_dimensions[col].width = 20
            boxSheet.column_dimensions[col].width = 20

        makeTubeLabel()

        excel.save(FILEPATH)



    

#excel.save(export_filename)

if __name__ == "__main__":
    initExcel()
    
    
'''

